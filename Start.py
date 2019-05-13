from parsing_modules import srx_parser, pan_parser, pan_panorama_parser, fortigate_parser, cisco_asa_parser
import argparse, sys, ipaddress, pprint, openpyxl

from openpyxl.styles import PatternFill, Font
from openpyxl.styles.colors import RED

from uuid import uuid4

import logging

pp = pprint.PrettyPrinter(indent=4)
Deny = Font(color=RED)

# create logger with 'spam_application'
logger = logging.getLogger('fwParser.log')
logger.setLevel(logging.CRITICAL)
# create file handler which logs even debug messages
fh = logging.FileHandler('fwParser.log')
fh.setLevel(logging.WARNING)
logger.addHandler(fh)

parser = argparse.ArgumentParser(prog='fwParser', allow_abbrev=True,
                                 description='Find matching policies in Firewall Configuration.')
parser.add_argument('vendor', metavar='<vendor>',
                    choices=['asa', 'fortigate', 'srx', 'pan-panorama', 'paloalto'],
                    help='Configuration vendor to be parsed (default: fortigate)\n'
                         'Supported (asa, fortigate, srx , panorama)')
parser.add_argument('filename', metavar='<configuration file>', help='Path to configuration file')

# parser.add_argument('-v', '--verbose', action='store_const', const=True, default=False, help='verbose mode')

parser.add_argument('--output', metavar='Filename', default='{}.xlsx'.format(uuid4()),
                    help='Output filename')

parser.add_argument('--subnet', metavar='subnet', nargs='*', default=["0.0.0.0/0"],
                    help='Target subnet(s) in X.X.X.X/X format')


def parse_vendor(argument):
    switcher = {
        'fortigate': fortigate_parser,
        'asa': cisco_asa_parser,
        'paloalto': pan_parser,
        'pan-panorama': pan_panorama_parser,
        'srx': srx_parser,
    }
    return switcher.get(argument, None)


target_config = None
results = dict()
if __name__ == '__main__':
    args = parser.parse_args()
    print(args)
    parser_engine = parse_vendor(args.vendor)
    if not parser_engine:
        print('No valid Engine to parse {} configuration file'.format(args.vendor))
    config = parser_engine(configfile=args.filename)
    config.parse()
    cfg = config.get_config_obj()
    if cfg.is_vdom():
        print('Found VDOM Configuration, Please select the target VDOM to continue:')
        for i, vdom in enumerate(cfg.get_vdoms()):
            print('{}) {}'.format(i, vdom))
        target_vdom = int(input('Enter the VDOM number:'))
        if target_vdom not in range(0, len(cfg.get_vdoms())):
            print('Invalid number please select number between {} and {}'.format('0', len(cfg.get_vdoms() - 1)))
            sys.exit(605)
        target_config = config.get_json_config()['vdom'][cfg.get_vdoms()[target_vdom]]
    else:
        target_config = config.get_json_config()

    # set result dict
    for arg in args.subnet:
        results[arg] = dict()
    print('Phase Alpha: Finding Matching Address Objects')
    for arg in results.keys():
        print('\n     Searching for Match for the subnet {}'.format(arg))
        results[arg]['AddrObj'] = dict()
        if 'firewall address' not in target_config.keys():
            print(
                'Provided configuration file does not contain \"Config Firewall Address Section\".\n '
                'Please ensure you have supplied valid configuration file.')
            sys.exit(6002)
        total = len(list(target_config['firewall address'].keys()))
        for i, addr in enumerate(target_config['firewall address'].keys()):
            matched = len(results[arg]['AddrObj'])
            sys.stdout.write(
                '\r     Processing {:<32} --- {:^6d}/{:^6d}/{:^6d} MATCHED/TESTED/TOTAL'.format(addr, matched, i + 1,
                                                                                                total))
            if 'type' not in target_config['firewall address'][addr]['set'].keys():
                target_config['firewall address'][addr]['set']['type'] = 'ipmask'
                if 'subnet' not in target_config['firewall address'][addr]['set'].keys():
                    target_config['firewall address'][addr]['set']['subnet'] = '0.0.0.0 0.0.0.0'
            if 'ipmask' is target_config['firewall address'][addr]['set']['type']:
                fwaddr = ipaddress.ip_network(
                    target_config['firewall address'][addr]['set']['subnet'].replace(' ', '/'))
                if ipaddress.ip_network(arg).overlaps(fwaddr):
                    results[arg]['AddrObj'][addr] = target_config['firewall address'][addr]['set']
                continue
            elif 'iprange' == target_config['firewall address'][addr]['set']['type']:
                hosts = 0
                start_addr = ipaddress.IPv4Address(target_config['firewall address'][addr]['set']['start-ip'])
                end_addr = ipaddress.IPv4Address(target_config['firewall address'][addr]['set']['end-ip'])
                if start_addr == end_addr:
                    if ipaddress.ip_network(arg).overlaps(ipaddress.ip_network(str(start_addr))):
                        results[arg]['AddrObj'][addr] = target_config['firewall address'][addr]['set']
                elif start_addr < end_addr:
                    while start_addr != end_addr + 1:
                        hosts += 1
                        if ipaddress.ip_network(arg).overlaps(ipaddress.ip_network(str(start_addr))):
                            results[arg]['AddrObj'][addr] = target_config['firewall address'][addr]['set']
                            break
                        start_addr += 1
                elif start_addr > end_addr:
                    count = 0
                    while end_addr != start_addr + 1:
                        if ipaddress.ip_network(arg).overlaps(ipaddress.ip_network(str(end_addr))):
                            results[arg]['AddrObj'][addr] = target_config['firewall address'][addr]['set']
                            break
                        end_addr += 1
                if hosts > 255:
                    logger.warning('\nObject {} contains more that {} /24 subnets'.format(addr, int(hosts / 255)))

    print('\nPhase Beta: Finding Matching Address Groups')
    for arg in results.keys():
        print('\n     Searching for Match for the subnet {}'.format(arg))
        results[arg]['AddrGrp'] = dict()
        for j, matched_addr_obj in enumerate(results[arg]['AddrObj'].keys()):
            total = len(target_config['firewall addrgrp'].keys())
            for i, grp in enumerate(target_config['firewall addrgrp'].keys()):
                matched = len(results[arg]['AddrGrp'].keys())
                sys.stdout.write(
                    '\r     Processing {:<32} --- {:^6d}/{:^6d}/{:^6d} MATCHED/TESTED/TOTAL'.format(grp, matched, i + 1,
                                                                                                    total))

                groups = target_config['firewall addrgrp'][grp]['set']['member'].split(' ')
                if matched_addr_obj in groups:
                    results[arg]['AddrGrp'][grp] = target_config['firewall addrgrp'][grp]['set']

    print('\nPhase gamma: Finding Matching Policies')
    for arg in results.keys():
        print('\n     Searching for Match for the subnet {}'.format(arg))
        results[arg]['Polcies'] = dict()
        for matched_addr_grp in results[arg]['AddrGrp'].keys():
            total = len(target_config['firewall policy'].keys())
            for i, policy in enumerate(target_config['firewall policy'].keys()):
                matched = len(results[arg]['Polcies'].keys())
                sys.stdout.write(
                    '\r     Processing Groups {:<32} --- {:^6d}/{:^6d}/{:^6d} MATCHED/TESTED/TOTAL'.format(policy,
                                                                                                           matched,
                                                                                                           i + 1,
                                                                                                           total))
                if 'srcaddr' not in target_config['firewall policy'][policy]['set'].keys():
                    target_config['firewall policy'][policy]['set']['srcaddr'] = '"all"'
                if 'dstaddr' not in target_config['firewall policy'][policy]['set'].keys():
                    target_config['firewall policy'][policy]['set']['dstaddr'] = '"all"'

                src = target_config['firewall policy'][policy]['set']['srcaddr'].split(' ')
                dst = target_config['firewall policy'][policy]['set']['dstaddr'].split(' ')
                if matched_addr_grp in src:
                    results[arg]['Polcies'][policy] = target_config['firewall policy'][policy]['set']
                elif matched_addr_grp in src:
                    results[arg]['Polcies'][policy] = target_config['firewall policy'][policy]['set']
        for matched_addr_obj in results[arg]['AddrObj'].keys():
            total = len(target_config['firewall policy'].keys())
            for i, policy in enumerate(target_config['firewall policy'].keys()):
                matched = len(results[arg]['Polcies'].keys())
                sys.stdout.write(
                    '\r     Processing Objects {:<32} --- {:^6d}/{:^6d}/{:^6d} MATCHED/TESTED/TOTAL'.format(policy,
                                                                                                            matched,
                                                                                                            i + 1,
                                                                                                            total))
                if 'srcaddr' not in target_config['firewall policy'][policy]['set'].keys():
                    target_config['firewall policy'][policy]['set']['srcaddr'] = '"all"'
                if 'dstaddr' not in target_config['firewall policy'][policy]['set'].keys():
                    target_config['firewall policy'][policy]['set']['dstaddr'] = '"all"'

                src = target_config['firewall policy'][policy]['set']['srcaddr'].split(' ')
                dst = target_config['firewall policy'][policy]['set']['dstaddr'].split(' ')
                if matched_addr_obj in src:
                    results[arg]['Polcies'][policy] = target_config['firewall policy'][policy]['set']
                elif matched_addr_obj in src:
                    results[arg]['Polcies'][policy] = target_config['firewall policy'][policy]['set']

    print('\nPhase delta: Cleaning the results')
    for arg in results.keys():
        results[arg]['ServicesObj'] = dict()
        results[arg]['ServicesGrp'] = dict()
        for policy in results[arg]['Polcies'].keys():
            services = results[arg]['Polcies'][policy]['service'].split(' ')
            for service in services:
                if service in target_config['firewall service custom'].keys():
                    if service not in results[arg]['ServicesObj'].keys():
                        results[arg]['ServicesObj'][service] = target_config['firewall service custom'][service]['set']
                if service in target_config['firewall service group'].keys():
                    results[arg]['ServicesGrp'][service] = target_config['firewall service group'][service]['set']
                    for srv in results[arg]['ServicesGrp'][service]['member'].split(' '):
                        if srv not in results[arg]['ServicesObj'].keys():
                            results[arg]['ServicesObj'][srv] = target_config['firewall service custom'][srv]['set']

    print('\nPhase epsilon: Export to Excel')
    # Set up the outpoot Workbook
    outxlsx = openpyxl.Workbook()
    outsheet = outxlsx.active
    outsheet.title = 'Policies'
    # Write the header
    outsheet['A1'] = 'Seq'
    outsheet['B1'] = 'Policy#'
    outsheet['C1'] = 'srcintf'
    outsheet['D1'] = 'dstintf'
    outsheet['E1'] = 'srcaddr'
    outsheet['F1'] = 'dstaddr'
    outsheet['G1'] = 'Service'
    outsheet['H1'] = 'Action'
    outsheet['I1'] = 'Status'
    outsheet['J1'] = 'comments'
    outsheet['K1'] = 'Targeted Subnet'
    row = 2
    for argCount, arg in enumerate(results.keys()):
        for seq, policy in enumerate(results[arg]['Polcies'].keys()):
            if 'status' not in results[arg]['Polcies'][policy].keys():
                results[arg]['Polcies'][policy]['status'] = 'enable'
            if 'comments' not in results[arg]['Polcies'][policy].keys():
                results[arg]['Polcies'][policy]['comments'] = ''
            if 'action' not in results[arg]['Polcies'][policy].keys():
                results[arg]['Polcies'][policy]['action'] = 'deny'

            # search for Address Groups
            groups_src = results[arg]['Polcies'][policy]['srcaddr'].split(" ")
            groups_dst = results[arg]['Polcies'][policy]['dstaddr'].split(" ")
            srcaddr = []
            dstaddr = []
            for grp_src in groups_src:
                if grp_src in target_config['firewall addrgrp'].keys():  # results[arg]['AddrGrp']: # firewall addrgrp
                    for member in target_config['firewall addrgrp'][grp_src]['set']['member'].split(' '):
                        srcaddr.append(member)
                else:
                    srcaddr.append(grp_src)
            for grp_dst in groups_dst:
                if grp_dst in target_config['firewall addrgrp'].keys():  # results[arg]['AddrGrp']: # firewall addrgrp
                    for member in target_config['firewall addrgrp'][grp_dst]['set']['member'].split(' '):
                        dstaddr.append(member)
                else:
                    dstaddr.append(grp_dst)

            scraddrRendered = []
            dstaddrRendered = []
            for src_obj in srcaddr:
                if src_obj in target_config['firewall address'].keys():
                    if 'ipmask' == target_config['firewall address'][src_obj]['set']['type']:
                        scraddrRendered.append(
                            str(ipaddress.ip_network(
                                target_config['firewall address'][src_obj]['set']['subnet'].replace(' ', '/'))))
                        continue
                    if "iprange" == target_config['firewall address'][src_obj]['set']['type']:
                        scraddrRendered.append('{}-{}'.format(
                            str(ipaddress.ip_network(target_config['firewall address'][src_obj]['set']['start-ip'])),
                            str(ipaddress.ip_network(target_config['firewall address'][src_obj]['set']['end-ip']))))
                        continue
                    print('Matching Only Supports Subnets and IP Range Objects: {}'.format(src_obj))

            for dst_obj in dstaddr:
                if dst_obj in target_config['firewall address'].keys():
                    if 'ipmask' == target_config['firewall address'][dst_obj]['set']['type']:
                        dstaddrRendered.append(
                            str(ipaddress.ip_network(
                                target_config['firewall address'][dst_obj]['set']['subnet'].replace(' ', '/'))))
                        continue
                    if "iprange" == target_config['firewall address'][dst_obj]['set']['type']:
                        dstaddrRendered.append('{}-{}'.format(
                            str(ipaddress.ip_network(target_config['firewall address'][dst_obj]['set']['start-ip'])),
                            str(ipaddress.ip_network(target_config['firewall address'][dst_obj]['set']['end-ip']))))
                        continue
                    input('Matching Only Supports Subnets and IP Range Objects: {}'.format(dst_obj))

            group_srv = results[arg]['Polcies'][policy]['service'].split(" ")
            service = []
            srv_data = []
            for grp_srv in group_srv:
                if grp_srv in results[arg]['ServicesGrp']:
                    for member in results[arg]['ServicesGrp'][grp_srv]['member'].split(' '):
                        service.append(member)
                else:
                    service.append(grp_srv)

            for srvObj in service:
                if 'tcp-portrange' in results[arg]['ServicesObj'][srvObj].keys():
                    srv_data.append('TCP/{}'.format(results[arg]['ServicesObj'][srvObj]['tcp-portrange']))
                if 'udp-portrange' in results[arg]['ServicesObj'][srvObj].keys():
                    srv_data.append('UDP/{}'.format(results[arg]['ServicesObj'][srvObj]['udp-portrange']))
                if 'protocol' in results[arg]['ServicesObj'][srvObj].keys():
                    srv_data.append('{}'.format(results[arg]['ServicesObj'][srvObj]['protocol']))
                if 'protocol' not in results[arg]['ServicesObj'][srvObj].keys() \
                        and 'tcp-portrange' not in results[arg]['ServicesObj'][srvObj].keys() \
                        and 'udp-portrange' not in results[arg]['ServicesObj'][srvObj].keys():
                    srv_data.append('Any Service')

            outsheet.cell(row=row, column=1).value = seq
            outsheet.cell(row=row, column=2).value = policy
            outsheet.cell(row=row, column=3).value = str(results[arg]['Polcies'][policy]['srcintf'].split(" "))
            outsheet.cell(row=row, column=4).value = str(results[arg]['Polcies'][policy]['dstintf'].split(" "))
            outsheet.cell(row=row, column=5).value = str(scraddrRendered)
            outsheet.cell(row=row, column=6).value = str(dstaddrRendered)
            outsheet.cell(row=row, column=7).value = str(srv_data)
            outsheet.cell(row=row, column=8).value = str(results[arg]['Polcies'][policy]['action'].split(" "))
            outsheet.cell(row=row, column=9).value = str(results[arg]['Polcies'][policy]['status'])
            outsheet.cell(row=row, column=10).value = str(target_config['firewall policy'][policy]['set']['comments'])
            outsheet.cell(row=row, column=11).value = str(arg)

            if 'deny' in results[arg]['Polcies'][policy]['action'].split(" "):
                for col in range(1, 12):
                    outsheet.cell(row=row, column=col).font = Deny
            if 'disable' in results[arg]['Polcies'][policy]['status'].split(" "):
                outsheet.cell(row=row, column=9).fill = PatternFill("solid", fgColor="FF0000")
            sys.stdout.write('\r  {:^6d} Policy Proccessed'.format(row - 1))
            row += 1

    print('\n----OPERATION COMPLETED----\n')

    try:
        outxlsx.save('{}'.format(args.output))
    except PermissionError as error:
        print("PermissionError: [Errno 13] Permission denied: {}".format(args.output))
        outxlsx.save('{}-{}.xlsx'.format(args.output, uuid4()))
